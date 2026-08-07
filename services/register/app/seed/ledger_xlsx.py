"""The Evam LEDGER workbook format — reading, vocabulary translation and (see
``build_ledger_workbook``) writing, in one self-contained module.

This is the FORMAT layer, deliberately independent of the import engine: everything
here is pure (openpyxl workbooks and plain dicts in, plain dicts and workbooks out —
no database session, no FastAPI), so the format can be tested, evolved and reused on
its own. ``app.seed.from_xlsx`` (the DB import engine) consumes the readers;
``app.api.export_ledger`` serves the writer.

Two workbook generations are understood:

* the v4 consolidated MIS (headers in row 1, flat 'Syndication' sheet), and
* the live "Evam Dashboard" ledger (banner/title rows, headers below; a two-section
  'Syndication Tracker'; 'Partnership Tracker'; Client/Lender/People masters) —
  exactly the workbook the desk maintains today.

The PRISM ledger EXPORT reproduces the Dashboard shape, so a file exported from PRISM
re-imports cleanly — and so does the desk's own historical file. Round-trip is the
contract.
"""

from __future__ import annotations

# --------------------------------------------------------------------------- #
# Reading: header detection, aliases, sections
# --------------------------------------------------------------------------- #

# Anchor headers that mark a REAL header row (lowercased). Any one of them is enough.
HDR_ANCHORS = {"client id", "lead id", "group code", "company name", "company",
               "company name (auto)", "company (auto)", "lender name", "role",
               "lender", "partner lender"}

# Ledger header spellings → the canonical header the parsers read.
HDR_ALIASES = {
    "company name (auto)": "Company Name",
    "company (auto)": "Company Name",
    "company": "Company Name",
    "mitigation / adaptation": "Mitigation / Adaptation",
    "updared remarks 19 july 2026": "Updated Remarks 19 July 2026",
    "sectors": "Sector",
    # The live ledger's Deals sheet spells the flag with a typo and a space.
    "partnerhship ?": "Partnership?",
    "partnership ?": "Partnership?",
}


def find_header_row(rows: list, limit: int = 6) -> int | None:
    """Index of the first row (within the top ``limit``) that looks like a header row."""
    for i, row in enumerate(rows[:limit]):
        cells = {" ".join(str(c).split()).lower() for c in row if c not in (None, "")}
        if len(cells) >= 2 and cells & HDR_ANCHORS:
            return i
    return 0 if rows else None


def rows_as_dicts(rows: list, hdr_ix: int, stop_at_banner: bool = False) -> list[dict]:
    """Rows below ``hdr_ix`` as dicts keyed by canonicalised, de-duplicated headers.

    A DUPLICATE header gets a positional suffix ('Status', 'Status#2') instead of
    silently shadowing the first — the ledger's Leads sheet carries both a lifecycle
    Status and a temperature Status. ``stop_at_banner`` ends the block at the next
    section banner (an ALL-CAPS cell in column A with nothing beside it) — the
    ledger's Syndication Tracker stacks two sections on one sheet."""
    raw_hdr = rows[hdr_ix]
    header: list[str] = []
    seen: dict[str, int] = {}
    for i, h in enumerate(raw_hdr):
        name = " ".join(str(h).split()) if h is not None else f"col{i}"
        name = HDR_ALIASES.get(name.lower(), name)
        n = seen.get(name, 0) + 1
        seen[name] = n
        header.append(name if n == 1 else f"{name}#{n}")
    out = []
    for r in rows[hdr_ix + 1:]:
        if not any(c not in (None, "") for c in r):
            continue
        first = str(r[0]).strip() if r and r[0] not in (None, "") else ""
        if stop_at_banner and first and first == first.upper() and len(first) > 12 \
                and not any(c not in (None, "") for c in r[1:]):
            break
        out.append({header[i]: r[i] for i in range(min(len(header), len(r)))})
    return out


def find_banner_row(rows: list, needle: str) -> int | None:
    """Index of the row whose first populated cell STARTS WITH ``needle`` (case-insensitive)."""
    for i, row in enumerate(rows):
        for c in row:
            if c in (None, ""):
                continue
            if str(c).strip().lower().startswith(needle.lower()):
                return i
            break
    return None


def sheet_rows(wb, title: str) -> list[dict]:
    """A sheet as a list of {header: value} dicts (non-empty rows only), reading both
    workbook generations — the header row is auto-detected."""
    if title not in wb.sheetnames:
        return []
    ws = wb[title]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    hdr_ix = find_header_row(rows)
    if hdr_ix is None:
        return []
    return rows_as_dicts(rows, hdr_ix)


def ledger_syn_rows(wb) -> list[dict]:
    """The ledger's 'Syndication Tracker' sheet, translated to the flat row shape the
    import engine consumes. The sheet stacks TWO sections: a derived deal-level view
    (skipped — every figure in it derives from the lender rows) and 'DETAILED
    LENDER-LEVEL ...' whose rows are the authoritative per-bank ledger."""
    if "Syndication Tracker" not in wb.sheetnames:
        return []
    rows = list(wb["Syndication Tracker"].iter_rows(values_only=True))
    banner = find_banner_row(rows, "DETAILED LENDER-LEVEL")
    if banner is None:
        return []
    out = []
    for r in rows_as_dicts(rows, banner + 1):
        out.append({
            "Client ID": r.get("Client ID"),
            "Company Name": r.get("Company Name"),
            "Bank": r.get("Lender"),
            "Status": r.get("Lender Status"),
            "Amount (₹ Cr)": r.get("Ticket Size (₹ Cr)"),
            # Same figure under its own name: the PER-LENDER ask. The engine puts it on
            # the lender row's allocation; "Amount (₹ Cr)" doubles as the tracker-level
            # fill (the v4 sheet's meaning). Absent from v4 rows on purpose.
            "Ticket Size (₹ Cr)": r.get("Ticket Size (₹ Cr)"),
            "RM": r.get("RM"),
            "Credit Analyst": r.get("Credit Analyst"),
            "Date Data Received": r.get("Date Data Received"),
            "Date IM Circulated": r.get("Date IM Circulated"),
            "Date In-Principle": r.get("Date In-Principle"),
            "Date Sanctioned": r.get("Date Sanctioned"),
            "Remarks": r.get("Remarks"),
            # PRISM-export carrier-row columns (absent from hand-kept ledgers): the
            # tracker's OWN status and ask, exact — so a Dropped mandate with live-
            # looking lender rows, or an ask differing from any single ticket, survives.
            "Tracker Status": r.get("Tracker Status"),
            "Tracker Ask (₹ Cr)": r.get("Tracker Ask (₹ Cr)"),
        })
    return out


# --------------------------------------------------------------------------- #
# Vocabulary translation (ledger wording ⇄ PRISM canonical)
# --------------------------------------------------------------------------- #

# Ledger per-LENDER statuses → PRISM's lender-pipeline vocabulary. The ledger says a
# bank "Rejected" (PRISM: Declined) and tracks lender-level "Disbursed" (PRISM keeps
# Disbursed at MANDATE level; the lender row lands Sanctioned and the original word is
# preserved on the note + reported as a translation).
LENDER_STATUS_MAP = {"rejected": "Declined", "disbursed": "Sanctioned"}
LENDER_VOCAB = {"Identified", "IM Circulated", "Docs Pending", "Queries Received",
                "IP Received", "Sanctioned", "Declined"}

# Forward order of the lender pipeline — a DUPLICATE ledger row for the same bank
# upgrades the row only when it sits further along ('Declined' ranks lowest so a live
# duplicate outranks a stale rejection; re-importing the identical file changes nothing).
LENDER_RANK = {"Declined": 1, "Identified": 2, "Docs Pending": 3, "IM Circulated": 4,
               "Queries Received": 5, "IP Received": 6, "Sanctioned": 7}


def canon_lender_status(value: str | None) -> tuple[str | None, bool]:
    """(canonical lender status, changed?) — case/space-insensitive against the lender
    vocabulary plus the ledger's Rejected/Disbursed wording."""
    if value is None:
        return None, False
    key = " ".join(str(value).split()).lower()
    mapped = LENDER_STATUS_MAP.get(key)
    if mapped is not None:
        return mapped, True
    for canonical in LENDER_VOCAB:
        if canonical.lower() == key:
            return canonical, canonical != value
    return value, False


# Ledger temperature spellings (typos included — they exist in the live data).
TEMP_CANON = {"hot": "Hot", "warm": "Warm", "wram": "Warm", "cold": "Cold"}


def canon_temp(value: str | None) -> tuple[str | None, bool]:
    if value is None:
        return None, False
    key = " ".join(str(value).split()).lower()
    out = TEMP_CANON.get(key)
    return (out, out != str(value).strip()) if out else (str(value).strip(), False)


# Lead lifecycle wording in the ledger.
LEAD_STATUS_CANON = {"active": "Active", "converted to deal": "Converted",
                     "converted": "Converted", "dropped": "Dropped", "on hold": "On Hold"}

# Mandate-pipeline rank, used to derive a tracker's status as the MOST ADVANCED of its
# lender rows (the ledger's own deal-level view derives the same way). 'Rejected' ranks
# lowest so a single live bank outranks any number of rejections — a mandate is
# Rejected only when EVERY bank declined.
SYN_RANK = {"Rejected": 1, "Dropped": 1, "Withdrawn": 1, "Deal Sourced": 2,
            "Docs Pending": 3, "IM in Prep": 4, "IM Circulated": 5,
            "Queries Received": 6, "IP Received": 7, "Sanctioned": 8, "Disbursed": 9}


# --------------------------------------------------------------------------- #
# Zero-loss helpers
# --------------------------------------------------------------------------- #

def extras_note(row: dict, used: set[str], skip: set[str] | None = None) -> str | None:
    """ZERO-LOSS backstop: every populated column the mapping did not consume is
    preserved as a readable '[Header: value]' tag (appended to the row's remarks/note),
    so a ledger column we never modelled — a colleague's notes column, a one-off
    remark header — survives import verbatim."""
    skip = skip or set()
    tags = []
    for h, v in row.items():
        if h in used or h in skip or v in (None, ""):
            continue
        base = h.split("#")[0]
        if base in used or base in skip or base.startswith("col"):
            continue
        sval = v.isoformat()[:10] if hasattr(v, "isoformat") else str(v).strip()
        if sval:
            tags.append(f"[{h}: {sval}]")
    return " ".join(tags) or None


def join_notes(*parts: str | None) -> str | None:
    return " | ".join(p for p in (x.strip() if x else None for x in parts) if p) or None


# PRISM tracker fields the Dashboard sheet has no column for. On EXPORT each populated
# field becomes a '[Label: value]' tag on the tracker's carrier row; on IMPORT the tags
# are parsed straight back into the fields — so a PRISM-entered facility/tenor/priority
# survives a full export → import cycle in its own field, not as loose text.
TRACKER_TAG_FIELDS = {
    "Facility": "facility", "Tenor": "tenor", "Priority": "priority",
    "Syndication type": "syndication_type", "IM status": "im_status",
    "Potential": "potential", "Existing": "existing", "Price": "price",
    "Mandate status detail": "mandate_status3", "Pending with": "pending_with",
    "TOI": "toi",
}

_TAG_RE = None  # compiled lazily (module import stays cheap)


def field_tags(rec: dict) -> str | None:
    """'[Label: value]' tags for every populated TRACKER_TAG_FIELDS field of ``rec``."""
    tags = []
    for label, field in TRACKER_TAG_FIELDS.items():
        v = rec.get(field)
        if v not in (None, ""):
            tags.append(f"[{label}: {v}]")
    return " ".join(tags) or None


def parse_field_tags(text: str | None) -> tuple[dict, str | None]:
    """(fields, remaining text) — the inverse of ``field_tags``. Unknown tags stay in
    the text untouched; only exact TRACKER_TAG_FIELDS labels are lifted out."""
    global _TAG_RE
    if not text:
        return {}, text
    if _TAG_RE is None:
        import re
        labels = "|".join(re.escape(lbl) for lbl in TRACKER_TAG_FIELDS)
        _TAG_RE = re.compile(r"\[(" + labels + r"): ([^\]]*)\]\s*")
    fields: dict = {}

    def _lift(m):
        fields[TRACKER_TAG_FIELDS[m.group(1)]] = m.group(2).strip() or None
        return ""

    rest = _TAG_RE.sub(_lift, text).strip(" |").strip()
    return fields, (rest or None)


def extract_tag(text: str | None, label: str) -> tuple[str | None, str | None]:
    """Lift a single '[label: value]' tag out of ``text`` → (value, remaining text)."""
    if not text:
        return None, text
    import re
    m = re.search(r"\[" + re.escape(label) + r": ([^\]]*)\]\s*", text)
    if not m:
        return None, text
    rest = (text[:m.start()] + text[m.end():]).strip(" |").strip()
    return m.group(1).strip() or None, (rest or None)


def split_mandate(mand: str | None) -> tuple[str | None, str | None, str | None, str | None]:
    """A stored mandate status back into the ledger's four columns:
    'Sent - Mandate Signed | [Syndication: Yes, Partnership: No]'
    → ('Sent', 'Mandate Signed', 'Yes', 'No'). Free-text that never came from the
    ledger lands whole in the first column — nothing is invented."""
    if not mand:
        return None, None, None, None
    syn = part = None
    base = mand
    if " | [" in mand:
        base, _, flag_part = mand.partition(" | [")
        flag_part = flag_part.rstrip("]")
        for piece in flag_part.split(","):
            k, _, v = piece.partition(":")
            k, v = k.strip().lower(), v.strip()
            if k == "syndication":
                syn = v
            elif k == "partnership":
                part = v
    sent, sep, signed = base.partition(" - ")
    return (sent.strip() or None, (signed.strip() or None) if sep else None, syn, part)


# --------------------------------------------------------------------------- #
# Writing: the PRISM ledger export (Dashboard-shaped, re-importable)
# --------------------------------------------------------------------------- #

# The ledger's own header rows, verbatim (the Deals sheet typo included — the reader
# aliases it, and the desk's eyes expect it).
_LEADS_HDR = ["Lead ID", "Source", "Source Detail", "RM Owner", "Status", "Sectors",
              "Mitigation / Adaptation", "Company Name", "Location", "Status",
              "Contact Person", "Designation", "Contact Phone", "Last Interaction Date",
              "Next Action", "Next Action Date", "Notes"]
_DEALS_HDR = ["Client ID", "Group Code", "Company Name (auto)", "Sector", "Location",
              "Source", "Source Detail", "Status", "RM", "Lending?", "Syndication?",
              "Partnerhship ?", "Asset Mon?", "Stage", "Date Received",
              "Contact Person", "Contact Phone", "Remarks"]
_LENDING_HDR = ["Client ID", "Company (auto)", "Lending Amount (₹ Cr)", "RM",
                "Credit Analyst", "Stage", "Stage Updated", "Pending With",
                "Date Allotted", "Date Initial Query Raised",
                "Date Client Reply Received", "Date Note Sent for Circulation",
                "Date Sanctioned",
                # PRISM columns the paper ledger never had — the importer reads these
                # exact names, so drawdown data round-trips in fields, not prose.
                "Proposed Disbursement Amount (₹ Cr)", "Proposed Disbursement Date",
                "Disbursed Amount (₹ Cr)", "Disbursement Date", "Remarks"]
_SYN_DEAL_HDR = ["Client ID", "Company", "Ticket Size (₹ Cr)", "Deal Status (derived)",
                 "Most Advanced Stage", "# Lenders"]
_SYN_LENDER_HDR = ["Client ID", "Company (auto)", "Ticket Size (₹ Cr)", "RM",
                   "Credit Analyst", "Lender", "Lender Status", "Date Data Received",
                   "Date IM Circulated", "Date In-Principle", "Date Sanctioned",
                   "Remarks",
                   # Carrier-row-only columns (see ledger_syn_rows): the tracker's own
                   # status and ask, kept exact across the round trip.
                   "Tracker Status", "Tracker Ask (₹ Cr)"]
_PART_HDR = ["Client ID", "Company (auto)", "RM", "Partner Lender", "Stage",
             "Stage Updated", "Pending With", "Sanctioned Amount (₹ Cr)",
             "Rejection Reason", "Remarks"]
_AM_HDR = ["Client ID", "Company (auto)", "RM", "Analyst", "State",
           "Indicative Value (₹ Cr)", "Size (MW)", "Nature", "Deal Type", "Investor",
           "Investor Type", "Status", "Date Teaser Shared", "Notes",
           "Updated Remarks 19 July 2026"]
_LENDER_MASTER_HDR = ["Lender Name", "Type", "Short Name", "Active?",
                      "Preferred Sectors", "Notes", "Total Submissions",
                      "Active Submissions", "Sanctioned", "Rejected"]
_CLIENT_MASTER_HDR = ["Group Code", "Company Legal Name", "Sector (default)",
                      "PAN (optional)", "Group Notes"]
_PEOPLE_HDR = ["Role", "Initials", "Full Name", "Notes"]
_MANDATE_HDR = ["Company", "RM", "Mandate Sent/Not Sent", "Signed/Pending",
                "Syndication", "Partnership"]

# Which milestone column a lender row's 'since' date belongs in, by current status.
_SINCE_COLUMN = {"Sanctioned": "Date Sanctioned", "IP Received": "Date In-Principle",
                 "IM Circulated": "Date IM Circulated", "Queries Received":
                 "Date IM Circulated"}


def _won_lost(status: str | None) -> str:
    if status in ("Sanctioned", "Disbursed"):
        return "Won"
    if status in ("Rejected", "Dropped", "Withdrawn"):
        return "Lost"
    return "Live"


def build_ledger_workbook(data: dict):
    """The PRISM register as a Dashboard-shaped ledger workbook — the desk's own sheet
    names, banner rows and headers, so the file reads like the ledger it replaces AND
    re-imports through ``app.seed.from_xlsx`` with nothing lost (that round trip is
    asserted by tests). Input is plain dicts keyed by model field names (no ORM, no
    session):

    ``entities, leads, deals, lending, syn_trackers, syn_lenders, am, counterparties,
    people`` (+ optional ``generated_at``/``tenant`` strings for the README).

    Client IDs are assigned EF-001… in the given entity order and used consistently
    across every sheet of the file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    bold = Font(bold=True)

    def _sheet_with(title: str, banner: str | None, header: list[str]):
        ws = wb.create_sheet(title)
        if banner:
            ws.append([banner])
            ws["A1"].font = bold
        ws.append(header)
        for cell in ws[ws.max_row]:
            cell.font = bold
        return ws

    entities = data.get("entities") or []
    cid: dict = {}       # entity_id -> EF-nnn
    ename: dict = {}     # entity_id -> legal name
    estate: dict = {}    # entity_id -> state (the Leads sheet's Location)
    for i, e in enumerate(entities):
        cid[e["id"]] = f"EF-{i + 1:03d}"
        ename[e["id"]] = e.get("legal_name")
        estate[e["id"]] = e.get("state")

    def _yn(v) -> str:
        return "Yes" if v else "No"

    # --- README (provenance) -------------------------------------------------
    ws = wb.active
    ws.title = "README"
    ws["A1"] = "EVAM DEAL FLOW MIS — exported from PRISM"
    ws["A1"].font = bold
    ws["A3"] = f"Generated: {data.get('generated_at') or ''}"
    ws["A4"] = f"Tenant: {data.get('tenant') or ''}"
    ws["A6"] = ("This file re-imports into PRISM (Admin → ledger import) with no data "
                "loss — statuses are PRISM's canonical vocabulary; details the sheet "
                "has no column for travel as [Tag: value] notes and are restored to "
                "their fields on import.")
    ws["A7"] = ("Client IDs (EF-nnn) are consistent within this file. The Dashboard / "
                "Reference Data sheets of the original ledger are live views; PRISM's "
                "own screens replace them, so they are not exported.")

    # --- Leads ---------------------------------------------------------------
    ws = _sheet_with("Leads", "LEADS", _LEADS_HDR)
    for ld in data.get("leads") or []:
        status = ld.get("status")
        ws.append([
            ld.get("lead_no"), ld.get("source"), ld.get("source_name"), ld.get("rm"),
            "Converted to Deal" if status == "Converted" else status,
            ld.get("sector"), ld.get("lens"), ld.get("company"),
            estate.get(ld.get("entity_id")), ld.get("temperature"), ld.get("contact"),
            ld.get("designation"), ld.get("phone"), ld.get("last_interaction_date"),
            ld.get("next_action"), ld.get("next_action_date"), ld.get("notes"),
        ])

    # --- Deals ---------------------------------------------------------------
    # Partnership? is derived from the company's Partnership-line tracker — the same
    # place the flag lands on import.
    partnership_entities = {t["entity_id"] for t in data.get("syn_trackers") or []
                            if t.get("line") == "Partnership"}
    ws = _sheet_with("Deals", "DEALS  —  one row per CLIENT (exported from PRISM)",
                     _DEALS_HDR)
    for d in data.get("deals") or []:
        eid_ = d.get("entity_id")
        ent = next((e for e in entities if e["id"] == eid_), {})
        ws.append([
            cid.get(eid_), d.get("code") or ent.get("code"), ename.get(eid_),
            ent.get("sector"), ent.get("state"), d.get("source"),
            d.get("source_detail"), d.get("temperature"), d.get("rm"),
            _yn(d.get("is_lending")), _yn(d.get("is_syndication")),
            _yn(eid_ in partnership_entities), _yn(d.get("is_asset_mon")),
            d.get("stage"), d.get("date_received"), None, None, d.get("remarks"),
        ])

    # --- Lending Tracker -----------------------------------------------------
    ws = _sheet_with("Lending Tracker", "LENDING TRACKER", _LENDING_HDR)
    for lt in data.get("lending") or []:
        eid_ = lt.get("entity_id")
        ws.append([
            cid.get(eid_), ename.get(eid_), lt.get("amount_cr"), lt.get("rm"),
            lt.get("analyst"), lt.get("stage"), lt.get("stage_updated_at"),
            lt.get("pending_with"), None, None, None, None, lt.get("sanction_date"),
            lt.get("proposed_disbursement_amount"), lt.get("proposed_disbursement_date"),
            lt.get("disbursed_amount"), lt.get("disbursement_date"), lt.get("remarks"),
        ])

    # --- Syndication Tracker (two sections, exactly like the ledger) ---------
    syn_trackers = [t for t in data.get("syn_trackers") or []
                    if t.get("line") != "Partnership"]
    part_trackers = [t for t in data.get("syn_trackers") or []
                     if t.get("line") == "Partnership"]
    lenders_by_tr: dict = {}
    for lr in data.get("syn_lenders") or []:
        lenders_by_tr.setdefault(lr["syndication_id"], []).append(lr)

    ws = wb.create_sheet("Syndication Tracker")
    ws.append(["SYNDICATION TRACKER"])
    ws["A1"].font = bold
    ws.append(["DEAL-LEVEL VIEW (derived from the lender rows below — the detailed "
               "section is the authoritative record)"])
    ws.append(_SYN_DEAL_HDR)
    for cell in ws[3]:
        cell.font = bold
    for t in syn_trackers:
        rows = lenders_by_tr.get(t["id"], [])
        ws.append([cid.get(t.get("entity_id")), ename.get(t.get("entity_id")),
                   t.get("amount_cr"), _won_lost(t.get("status")), t.get("status"),
                   len(rows)])
    ws.append([])
    ws.append(["DETAILED LENDER-LEVEL ROWS"])
    ws[ws.max_row][0].font = bold
    ws.append(_SYN_LENDER_HDR)
    for cell in ws[ws.max_row]:
        cell.font = bold

    def _lender_dates(lr: dict) -> dict:
        """since placed in the milestone column its status names (the importer reads it
        back to the same 'since'); any other original dates live on in the note tags."""
        out = {h: None for h in ("Date Data Received", "Date IM Circulated",
                                 "Date In-Principle", "Date Sanctioned")}
        col = _SINCE_COLUMN.get(lr.get("status") or "", "Date Data Received")
        out[col] = lr.get("since")
        return out

    for t in syn_trackers:
        tcid, tnm = cid.get(t.get("entity_id")), ename.get(t.get("entity_id"))
        rows = lenders_by_tr.get(t["id"], [])
        # The tracker's CARRIER row: its own status and ask (exact, in their own
        # columns), its remarks, and a tag for every PRISM field the sheet has no
        # column for. The importer applies it after the lender rows, so a tracker
        # whose status outranks (or trails) its lender rows survives verbatim.
        ws.append([tcid, tnm, None, t.get("rm"), t.get("analyst"), None, None,
                   None, None, None, None,
                   join_notes(field_tags(t), t.get("remarks")),
                   t.get("status"), t.get("amount_cr")])
        for lr in rows:
            dts = _lender_dates(lr)
            ws.append([tcid, tnm, lr.get("amount_cr"), t.get("rm"), t.get("analyst"),
                       lr.get("lender_name"), lr.get("status"),
                       dts["Date Data Received"], dts["Date IM Circulated"],
                       dts["Date In-Principle"], dts["Date Sanctioned"],
                       lr.get("note")])

    # --- Partnership Tracker -------------------------------------------------
    ws = _sheet_with("Partnership Tracker", "PARTNERSHIP TRACKER", _PART_HDR)
    for t in part_trackers:
        tcid, tnm = cid.get(t.get("entity_id")), ename.get(t.get("entity_id"))
        rows = lenders_by_tr.get(t["id"], [])
        tags = field_tags({k: v for k, v in t.items() if k != "pending_with"})
        # Carrier row (blank Partner Lender): the tracker's own stage, pending-with
        # and remarks — always emitted, so the tracker round-trips exactly.
        ws.append([tcid, tnm, t.get("rm"), None, t.get("status"), None,
                   t.get("pending_with"), None, None,
                   join_notes(tags, t.get("remarks"))])
        for lr in rows:
            reason, rest = extract_tag(lr.get("note"), "Rejection reason")
            ws.append([tcid, tnm, t.get("rm"), lr.get("lender_name"), lr.get("status"),
                       lr.get("since"), t.get("pending_with"), lr.get("amount_cr"),
                       reason, rest])

    # --- Asset Mon Tracker ---------------------------------------------------
    ws = _sheet_with("Asset Mon Tracker", "ASSET MONETISATION TRACKER", _AM_HDR)
    for a in data.get("am") or []:
        eid_ = a.get("entity_id")
        ws.append([
            cid.get(eid_), ename.get(eid_), a.get("rm"), a.get("analyst"),
            a.get("state"), a.get("indicative_value_cr"), a.get("size_mw"),
            a.get("nature"), a.get("deal_type"), a.get("investor"),
            a.get("investor_type"), a.get("status"), a.get("teaser_date"),
            a.get("notes"), None,
        ])

    # --- Lender Master (with derived submission stats, like the ledger's) ----
    subs: dict = {}
    for lr in data.get("syn_lenders") or []:
        k = (lr.get("lender_name") or "").lower()
        s = subs.setdefault(k, {"total": 0, "active": 0, "won": 0, "lost": 0})
        s["total"] += 1
        st = lr.get("status")
        if st == "Sanctioned":
            s["won"] += 1
        elif st == "Declined":
            s["lost"] += 1
        else:
            s["active"] += 1
    ws = _sheet_with("Lender Master", "LENDER MASTER", _LENDER_MASTER_HDR)
    for c in data.get("counterparties") or []:
        s = subs.get((c.get("name") or "").lower(), {})
        ws.append([c.get("name"), c.get("counterparty_type"), c.get("short_name"),
                   _yn(c.get("is_active", True)), c.get("sectors"), c.get("notes"),
                   s.get("total"), s.get("active"), s.get("won"), s.get("lost")])

    # --- Client Master -------------------------------------------------------
    ws = _sheet_with("Client Master",
                     "CLIENT MASTER  —  canonical company registry", _CLIENT_MASTER_HDR)
    for e in entities:
        ws.append([e.get("code"), e.get("legal_name"), e.get("sector"),
                   e.get("pan"), e.get("notes")])

    # --- People Master -------------------------------------------------------
    ws = _sheet_with("People Master", "PEOPLE MASTER", _PEOPLE_HDR)
    for p in data.get("people") or []:
        ws.append([p.get("role"), p.get("name"), p.get("full_name"), p.get("notes")])

    # --- Mandate Tracker (header in row 1, exactly like the ledger) ----------
    ws = wb.create_sheet("Mandate Tracker")
    ws.append(_MANDATE_HDR)
    for cell in ws[1]:
        cell.font = bold
    for t in (data.get("syn_trackers") or []):
        if not t.get("mandate_status"):
            continue
        sent, signed, syn_flag, part_flag = split_mandate(t.get("mandate_status"))
        ws.append([ename.get(t.get("entity_id")), t.get("rm"), sent, signed,
                   syn_flag, part_flag])

    return wb
