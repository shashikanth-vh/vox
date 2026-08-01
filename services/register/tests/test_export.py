"""DB → Excel / JSON export."""

from __future__ import annotations

import io

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook

pytestmark = pytest.mark.asyncio


async def _seed(client: AsyncClient) -> None:
    eid = (await client.post("/v1/entities", json={"code": "EXP1", "legal_name": "Export Co"})).json()["id"]
    await client.post("/v1/leads", json={"company": "Lead Co"})
    await client.post(f"/v1/entities/{eid}/interactions",
                      json={"interaction_type": "Phone Call", "notes": "hi"})


async def test_export_counts(client: AsyncClient):
    await _seed(client)
    r = await client.get("/v1/export/counts")
    assert r.status_code == 200
    c = r.json()
    assert c["entities"] == 1 and c["leads"] == 1 and c["interactions"] == 1


async def test_export_excel(client: AsyncClient):
    await _seed(client)
    r = await client.get("/v1/export/excel")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert r.headers["content-disposition"].endswith('.xlsx"')

    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    # One sheet per table.
    assert {"entities", "leads", "interactions", "deals"} <= set(wb.sheetnames)
    ws = wb["entities"]
    rows = list(ws.iter_rows(values_only=True))
    header, data = rows[0], rows[1:]
    assert "legal_name" in header and "id" in header and "version" in header
    assert len(data) == 1
    assert data[0][header.index("legal_name")] == "Export Co"


async def test_export_json_and_subset(client: AsyncClient):
    await _seed(client)
    r = await client.get("/v1/export/json", params={"tables": "entities,leads"})
    assert r.status_code == 200
    body = r.json()
    assert set(body["tables"].keys()) == {"entities", "leads"}
    assert body["tables"]["entities"][0]["legal_name"] == "Export Co"
    assert "exported_at" in body and body["tenant"] == "EVAM"


async def test_export_excel_subset(client: AsyncClient):
    await _seed(client)
    r = await client.get("/v1/export/excel", params={"tables": "leads"})
    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    assert wb.sheetnames == ["leads"]


async def test_export_carries_the_release1_operational_registers(client: AsyncClient):
    """The Excel tracker's activities all live in PRISM now — so the workbook must carry
    every Release-1 register: calendar, covenants + observations, EWS, governance
    evidence, decisions, CP/CS, handover, tranches and notifications."""
    await _seed(client)
    # One row in each new register (minimum viable — the journey collection fills them all).
    eid = (await client.get("/v1/entities", params={"q": "Export Co"})).json()["items"][0]["id"]
    ch = {"X-User-Email": "ch@evamfinance.com", "X-User-Roles": "Credit Head"}
    cal = await client.post("/v1/calendar-events",
                            json={"title": "Review", "starts_at": "2026-09-01T09:00:00Z"},
                            headers=ch)
    assert cal.status_code == 201, cal.text
    cov = await client.post("/v1/covenants", json={
        "entity_id": eid, "name": "DSCR >= 1.2", "covenant_type": "Financial",
        "metric": "dscr", "operator": ">=", "threshold": 1.2,
        "frequency": "Quarterly", "first_due_on": "2026-01-01"}, headers=ch)
    assert cov.status_code == 201, cov.text
    case = await client.post("/v1/ews-cases", json={
        "entity_id": eid, "title": "Watch", "source": "manual"}, headers=ch)
    assert case.status_code == 201, case.text

    counts = (await client.get("/v1/export/counts")).json()
    m = counts.get("counts", counts)
    for t in ("calendar_events", "covenants", "ews_cases"):
        assert m.get(t, 0) >= 1, f"{t} missing from counts: {m}"
    # Every operational register appears as a sheet in the workbook.
    r = await client.get("/v1/export/excel")
    assert r.status_code == 200, r.text
    wb = load_workbook(io.BytesIO(r.content), read_only=True)
    assert {"calendar_events", "covenants", "ews_cases", "governance_evidence",
            "workflow_decisions", "cp_cs_checklists", "advaya_handover_packages",
            "disbursement_tranches", "notifications"} <= set(wb.sheetnames)
    ws = wb["covenants"]
    header = next(ws.iter_rows(values_only=True))
    assert "threshold" in header and "frequency" in header
