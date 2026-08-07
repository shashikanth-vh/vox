"""The ledger round trip — the Dashboard-shaped workbook in, the same substance out.

The contract under test: a file shaped like the desk's live 'Evam Dashboard' ledger
(title rows, headers below, dual Status columns on Leads, a two-section Syndication
Tracker, Partnership Tracker, the three masters, legacy vocabulary like
Rejected / Disbursed / 'Wram' / 'Converted to Deal') imports with nothing lost; the
ledger export reproduces that shape; and importing the export changes NOTHING — the
same counts, the same statuses, no note growth. Seamless forever, per the desk's ask.
"""

from __future__ import annotations

import io

import pytest
from httpx import AsyncClient
from openpyxl import Workbook

pytestmark = pytest.mark.asyncio

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _dashboard_ledger() -> bytes:
    """A miniature of the REAL ledger workbook: same sheet names, same banner/title
    rows, same headers (typos included), same legacy wording."""
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Leads")
    ws.append(["LEADS"])                        # title row above the header, like the file
    ws.append(["Lead ID", "Source", "Source Detail", "RM Owner", "Status", "Sectors",
               "Mitigation / Adaptation", "Company Name", "Location", "Status",
               "Contact Person", "Designation", "Contact Phone", "Last Interaction Date",
               "Next Action", "Next Action Date", "Notes"])
    # dual Status: lifecycle first, temperature second (with the live file's typo)
    ws.append(["LD-001", "RM", "Shubh", "SD", "Active", "Solar", "Mitigation",
               "Mir Green Energy", "KA", "Wram", "Pramod", "", "98453", None,
               "Follow up", None, "note-1"])
    ws.append(["LD-002", "RM", "Shubh", "SD", "Converted to Deal", "Solar", "Mitigation",
               "Kahn Mecpower Energy Private Limited", "GJ", "Hot", "", "", "", None,
               "", None, ""])
    ws.append([None, "RM", "Chetan", "CM", "Active", "Solar", "Mitigation",
               None, "TN", "cold", "", "", "", None, "", None, "no company on this row"])

    ws = wb.create_sheet("Deals")
    ws.append(["DEALS  —  one row per CLIENT"])
    ws.append(["Client ID", "Group Code", "Company Name (auto)", "Sector", "Location",
               "Source", "Source Detail", "Status", "RM", "Lending?", "Syndication?",
               "Partnerhship ?", "Asset Mon?", "Stage", "Date Received",
               "Contact Person", "Contact Phone", "Remarks"])
    ws.append(["EF-001", "KAHNMEC", "Kahn Mecpower Energy Private Limited",
               "Renewables - Solar", "GJ", "RM", "", "Cold", "CM", "Yes", "Yes", "No",
               "No", "Closed Won", "2025-12-15", "", "", "sanctioned and moving"])
    ws.append(["EF-002", "AESPL", "Aapaavani Environmental Solutions Private Limited",
               "Water & Waste Management", "KA", "DSA", "Mithun", "Hot", "SD", "Yes",
               "No", "Yes", "No", "In Pipeline", "2026-01-07", "", "", ""])

    ws = wb.create_sheet("Lending Tracker")
    ws.append(["LENDING TRACKER"])
    ws.append(["Client ID", "Company (auto)", "Lending Amount (₹ Cr)", "RM",
               "Credit Analyst", "Stage", "Stage Updated", "Pending With",
               "Date Allotted", "Date Initial Query Raised", "Date Client Reply Received",
               "Date Note Sent for Circulation", "Date Sanctioned", "Remarks"])
    # ledger wording 'Disbursed' + 'Rejected'; Pending With must survive (it used not to)
    ws.append(["EF-002", "Aapaavani Environmental Solutions Private Limited", 1.0, "SD",
               "AT", "Disbursed", "2026-06-03", "Client", None, None, None, None, None,
               "Disbursed, Rs 50 Lakh first tranche"])
    ws.append(["EF-001", "Kahn Mecpower Energy Private Limited", 2.0, "SD", "AT",
               "Rejected", "2026-05-05", None, None, None, None, None, None, "weak coverage"])

    ws = wb.create_sheet("Syndication Tracker")
    ws.append(["SYNDICATION TRACKER"])
    ws.append(["DEAL-LEVEL VIEW (auto-derived)"])
    ws.append(["Client ID", "Company", "Ticket Size (₹ Cr)", "Deal Status (derived)",
               "Most Advanced Stage", "# Lenders"])
    ws.append(["EF-001", "Kahn Mecpower Energy Private Limited", 9.75, "Won",
               "Sanctioned", 2])                # derived section: skipped by the importer
    ws.append([])
    ws.append(["DETAILED LENDER-LEVEL ROWS"])
    ws.append(["Client ID", "Company (auto)", "Ticket Size (₹ Cr)", "RM",
               "Credit Analyst", "Lender", "Lender Status", "Date Data Received",
               "Date IM Circulated", "Date In-Principle", "Date Sanctioned", "Remarks"])
    ws.append(["EF-001", "Kahn Mecpower Energy Private Limited", 9.75, "CM", "AT",
               "Axis Bank", "Sanctioned", "2026-03-15", "2026-04-10", None, "2026-05-01",
               "sanctioned, docs WIP"])
    ws.append(["EF-001", "Kahn Mecpower Energy Private Limited", 5, "CM", "AT",
               "Kotak Mahindra", "Rejected", None, "2026-04-08", None, None, "not keen"])
    # same bank twice on one company (real in the live file) → must MERGE, not vanish
    ws.append(["EF-001", "Kahn Mecpower Energy Private Limited", 3, "CM", "AT",
               "Axis Bank", "IM Circulated", None, "2026-04-12", None, None,
               "second facility"])
    # a row with no lender named yet → substance lands on the tracker, not dropped
    ws.append(["EF-001", "Kahn Mecpower Energy Private Limited", None, "CM", "AT",
               None, None, "2026-03-01", None, None, None, "data pack received"])

    ws = wb.create_sheet("Partnership Tracker")
    ws.append(["PARTNERSHIP TRACKER"])
    ws.append(["Client ID", "Company (auto)", "RM", "Partner Lender", "Stage",
               "Stage Updated", "Pending With", "Sanctioned Amount (₹ Cr)",
               "Rejection Reason", "Remarks"])
    ws.append(["EF-002", "Aapaavani Environmental Solutions Private Limited", "SD",
               "Others", "IM Circulated", "2026-05-14", "RM", None, None,
               "IM circulated to Orix"])
    # blank partner: the tracker still forms; the remark survives on it
    ws.append(["EF-002", "Aapaavani Environmental Solutions Private Limited", "SD",
               None, "Docs Pending", "2026-05-14", "RM", None, None, "Data present"])

    ws = wb.create_sheet("Asset Mon Tracker")
    ws.append(["ASSET MONETISATION TRACKER"])
    ws.append(["Client ID", "Company (auto)", "RM", "State", "Indicative Value (₹ Cr)",
               "Size (MW)", "Nature", "Deal Type", "Investor", "Investor Type", "Status",
               "Date Teaser Shared", "Notes", "Updared Remarks 19 July 2026"])
    ws.append(["EF-003", "Axel Renewable Private Limited", "", "MH", 270.0, 58.5,
               "Seller", "Capital Market", "Exovolt", "VC", "In Discussion",
               "2026-04-16", "Solar + BESS", "45 MW under discussion"])

    ws = wb.create_sheet("Lender Master")
    ws.append(["LENDER MASTER"])
    ws.append(["Lender Name", "Type", "Short Name", "Active?", "Preferred Sectors",
               "Notes"])
    ws.append(["Axis Bank", "Bank", "Axis", "Yes", "Solar, MSME", ""])
    ws.append(["Kotak Mahindra", "Bank", "Kotak", "Yes", "Solar, BESS", ""])

    ws = wb.create_sheet("Client Master")
    ws.append(["CLIENT MASTER  —  canonical registry"])
    ws.append(["Group Code", "Company Legal Name", "Sector (default)", "PAN (optional)",
               "Group Notes"])
    ws.append(["KAHNMEC", "Kahn Mecpower Energy Private Limited", "Renewables - Solar",
               "AAACK1234A", ""])
    ws.append(["AESPL", "Aapaavani Environmental Solutions Private Limited",
               "Water & Waste Management", "", ""])

    ws = wb.create_sheet("People Master")
    ws.append(["PEOPLE MASTER"])
    ws.append(["Role", "Initials", "Full Name", "Notes"])
    ws.append(["RM", "SD", "Shubh Dave", ""])
    ws.append(["RM", "CM", "Chetan Malik", ""])
    ws.append(["Credit Analyst", "AT", "Archana Tripathi", ""])

    ws = wb.create_sheet("Mandate Tracker")
    ws.append(["Company", "RM", "Mandate Sent/Not Sent", "Signed/Pending",
               "Syndication", "Partnership"])
    ws.append(["Kahn Mecpower Energy Private Limited", "Shubh Dave", "Sent",
               "Mandate Signed", "Yes", "No"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def _import(client: AsyncClient, blob: bytes, mode: str) -> dict:
    r = await client.post("/v1/import/atlas-xlsx",
                          params={"mode": mode, "reason": "ledger round-trip test",
                                  "retain_incomplete": True},
                          files={"file": ("ledger.xlsx", blob, XLSX_MIME)})
    assert r.status_code == 200, r.text
    return r.json()


async def _state(client: AsyncClient) -> dict:
    ents = (await client.get("/v1/entities", params={"with_total": True})).json()["total"]
    leads = (await client.get("/v1/leads", params={"with_total": True})).json()["total"]
    deals = (await client.get("/v1/deals", params={"with_total": True})).json()["total"]
    syn = (await client.get("/v1/syndication", params={"with_total": True})).json()["items"]
    lend = (await client.get("/v1/lending", params={"with_total": True})).json()["items"]
    am = (await client.get("/v1/asset-monetisation",
                           params={"with_total": True})).json()["total"]
    lenders: dict[str, dict] = {}
    for t in syn:
        for ln in (await client.get(f"/v1/syndication/{t['id']}/lenders")).json():
            key = f"{t.get('line') or 'Syndication'}:{ln['lender_name']}"
            lenders[key] = {"status": ln.get("status"), "amount": ln.get("amount_cr"),
                            "note": ln.get("note")}
    return {
        "entities": ents, "leads": leads, "deals": deals, "am": am,
        "lending": sorted((x.get("stage") or "", x.get("pending_with") or "",
                           x.get("remarks") or "") for x in lend),
        "trackers": sorted((t.get("line") or "", t.get("status") or "",
                            t.get("mandate_status") or "", t.get("remarks") or "")
                           for t in syn),
        "lenders": lenders,
    }


async def test_dashboard_ledger_imports_with_zero_loss(client: AsyncClient):
    out = await _import(client, _dashboard_ledger(), "replace")
    c, rep = out["counts"], out["report"]

    assert c["leads"] == 2                       # the blank-company row is not a lead...
    q = rep["quarantined"]
    assert any(x["sheet"] == "Leads" and x["reason"] == "row has no company name"
               and "no company on this row" in (x.get("value") or "") for x in q), q

    # dual Status: lifecycle + temperature, both canonicalised with a record
    r = await client.get("/v1/leads", params={"q": "Mir Green"})
    lead = next(x for x in r.json()["items"] if "Mir Green" in x["company"])
    assert lead["temperature"] == "Warm"         # 'Wram' healed
    r = await client.get("/v1/leads", params={"q": "Kahn"})
    kahn_lead = next(x for x in r.json()["items"] if "Kahn" in x["company"])
    assert kahn_lead["status"] == "Converted"
    assert kahn_lead["converted_deal_id"] is not None

    # lending: ledger wording mapped, Pending With preserved
    lend = (await client.get("/v1/lending")).json()["items"]
    aespl = next(x for x in lend if x.get("pending_with") == "Client")
    assert aespl["stage"] == "Disbursed"

    # syndication: 2 named lenders (the duplicate Axis row MERGED, recorded)
    assert c["syndication_lenders"] == 2
    assert any("duplicate row for lender 'Axis Bank'" in (d.get("note") or "")
               for d in rep["derived"])
    syn = (await client.get("/v1/syndication")).json()["items"]
    mand = next(t for t in syn if (t.get("line") or "Syndication") != "Partnership")
    rows = (await client.get(f"/v1/syndication/{mand['id']}/lenders")).json()
    axis = next(x for x in rows if x["lender_name"] == "Axis Bank")
    assert axis["status"] == "Sanctioned"        # the duplicate did not regress it
    assert "second facility" in (axis["note"] or "")
    # the no-lender row's substance survives on the tracker
    assert "data pack received" in (mand.get("remarks") or "")

    # partnership: tracker formed; blank-partner remark preserved on it
    assert c["partnership_tracker"] == 1 and c["partnership_lenders"] == 1
    part = next(t for t in syn if t.get("line") == "Partnership")
    assert "Data present" in (part.get("remarks") or "")

    # the deal flags reconcile with the partnership tracker
    deals = (await client.get("/v1/deals")).json()["items"]
    aespl_deal = next(d for d in deals if d.get("code") == "AESPL")
    assert aespl_deal["is_syndication"] is True

    # masters: Client Master code became the entity code
    ents = (await client.get("/v1/entities")).json()["items"]
    assert {e["code"] for e in ents} >= {"KAHNMEC", "AESPL"}


async def test_export_reimports_identically(client: AsyncClient):
    """import → export → import must be a FIXED POINT: same counts, same statuses,
    same notes — and a further merge re-import changes nothing at all."""
    await _import(client, _dashboard_ledger(), "replace")
    before = await _state(client)

    r = await client.get("/v1/export/ledger-xlsx")
    assert r.status_code == 200, r.text
    assert r.headers["content-type"].startswith("application/vnd.openxmlformats")
    exported = r.content

    out = await _import(client, exported, "replace")
    assert out["report"]["quarantined_count"] == 0, out["report"]["quarantined"]
    after = await _state(client)
    assert after == before

    # merge re-import of the same export: a strict no-op
    out = await _import(client, exported, "merge")
    c = out["counts"]
    assert c["entities"] == 0 and c["leads"] == 0 and c["deals"] == 0
    assert c["syndication_tracker"] == 0 and c["syndication_lenders"] == 0
    assert c["partnership_tracker"] == 0 and c["partnership_lenders"] == 0
    assert await _state(client) == after


async def test_prism_only_vocabulary_survives_the_trip(client: AsyncClient):
    """Statuses that exist only in PRISM's lender pipeline ('Identified') and tracker
    facts with no ledger column (facility, tenor) must survive export → import in
    their own fields — not as free text, and never quarantined."""
    await _import(client, _dashboard_ledger(), "replace")
    syn = (await client.get("/v1/syndication")).json()["items"]
    mand = next(t for t in syn if (t.get("line") or "Syndication") != "Partnership")

    # PRISM-side edits: a freshly identified bank; tracker facts with no ledger column
    r = await client.post(f"/v1/syndication/{mand['id']}/lenders",
                          json={"lender_name": "HDFC Bank", "status": "Identified"})
    assert r.status_code == 201, r.text
    r = await client.patch(f"/v1/syndication/{mand['id']}",
                           json={"facility": "Term Loan", "tenor": "7y"},
                           headers={"If-Match": f'"{mand["version"]}"'})
    assert r.status_code == 200, r.text

    exported = (await client.get("/v1/export/ledger-xlsx")).content
    out = await _import(client, exported, "replace")
    assert out["report"]["quarantined_count"] == 0, out["report"]["quarantined"]

    syn2 = (await client.get("/v1/syndication")).json()["items"]
    mand2 = next(t for t in syn2 if (t.get("line") or "Syndication") != "Partnership")
    assert mand2["facility"] == "Term Loan" and mand2["tenor"] == "7y"
    rows2 = (await client.get(f"/v1/syndication/{mand2['id']}/lenders")).json()
    hdfc = next(x for x in rows2 if x["lender_name"] == "HDFC Bank")
    assert hdfc["status"] == "Identified"
