"""ATLAS MIS xlsx import (upload endpoint)."""

from __future__ import annotations

import io

import pytest
from httpx import AsyncClient
from openpyxl import Workbook

pytestmark = pytest.mark.asyncio


def _mini_mis() -> bytes:
    """A tiny 6-sheet workbook shaped like the real MIS."""
    wb = Workbook()
    wb.remove(wb.active)

    ws = wb.create_sheet("Leads")
    ws.append(["Source", "Source Detail", "RM Owner", "Sector", "Mitigation / Adaptation",
               "Company Name", "Status", "Contact Person", "Designation", "Contact Phone",
               "Last Interaction Date", "Next Action", "Next Action Date", "Notes"])
    ws.append(["RM", "Shubh", "Shubh Dave", "Renewables - Solar", "Mitigation", "EcoSoch",
               "Cold", "Harsha", "", "9535240349", None, "Follow up", None, "note"])

    ws = wb.create_sheet("Deals")
    ws.append(["Company Name", "Sector", "Location", "Source", "Source Detail", "Status", "RM",
               "Lending?", "Syndication?", "Asset Mon?", "Stage", "Date Received",
               "Contact Person", "Contact Phone", "Remarks"])
    ws.append(["ANV Web Ventures Private Limited", "EV", "TG", "DSA", "Get Vantage", "Hot",
               "Shubh Dave", "Yes", "Yes", "No", "In Pipeline", None, "", "", "in process"])

    ws = wb.create_sheet("Lending Tracker")
    ws.append(["Company Name", "Lending Amount (₹ Cr)", "RM", "Credit Analyst", "Stage",
               "Stage Updated", "Remarks", "Proposed Disbursement Amount (₹ Cr)",
               "Proposed Disbursement Date"])
    # A 'Ready for Disbursement' line carries its mandatory amount + date (the default importer
    # quarantines such a row missing them, exactly as the interactive API rejects it).
    ws.append(["ANV Web Ventures Private Limited", 1.15, "Shubh Dave", "AT",
               "Ready for Disbursement", None, "ok", 1.15, "2026-01-10"])

    ws = wb.create_sheet("Syndication")
    ws.append(["Company Name", "Deal Status", "Bank", "Status", "Amount (₹ Cr)",
               "Accepted by Client", "Remarks"])
    ws.append(["ANV Web Ventures Private Limited", "IM Circulated", "Axis Finance", "IM Circulated",
               10, "No", "shared"])
    ws.append(["ANV Web Ventures Private Limited", "IM Circulated", "Bajaj Finance", "Rejected",
               10, "No", "dropped"])

    ws = wb.create_sheet("Asset Mon")
    ws.append(["Company Name", "RM", "State", "Indicative Value (₹ Cr)", "Size (MW)", "Nature",
               "Deal Type", "Investor", "Investor Type", "Status", "Date Teaser Shared",
               "Notes", "Updated Remarks 19 July 2026"])
    ws.append(["Axel Renewable Private Limited", "Shubh Dave", "MH", 270, 58, "Seller",
               "Capital Market", "Radiance", "VC", "In Discussion", None, "solar+bess", "upd"])

    ws = wb.create_sheet("Mandate Tracker")
    ws.append(["Company Name", "RM", "Mandate Sent/Not Sent", "Signed/Pending"])
    ws.append(["ANV Web Ventures Private Limited", "Shubh Dave", "Sent", "Mandate Signed"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def test_import_atlas_xlsx_replace(client: AsyncClient):
    files = {"file": ("mis.xlsx", _mini_mis(),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = await client.post("/v1/import/atlas-xlsx",
                          params={"mode": "replace", "reason": "test import"}, files=files)
    assert r.status_code == 200, r.text
    counts = r.json()["counts"]
    # 3 distinct companies across the sheets → 3 entities
    assert counts["entities"] == 3
    assert counts["leads"] == 1
    assert counts["deals"] == 1
    assert counts["lending_tracker"] == 1
    assert counts["syndication_tracker"] == 1
    assert counts["syndication_lenders"] == 2
    assert counts["asset_monetisation"] == 1
    assert counts["mandate_applied"] == 1

    # entities landed and are queryable
    r = await client.get("/v1/entities", params={"with_total": True})
    assert r.json()["total"] == 3

    # the mandate rolled onto the company's syndication tracker
    r = await client.get("/v1/syndication", params={"with_total": True})
    tr = r.json()["items"][0]
    assert tr["mandate_status"] == "Sent - Mandate Signed"

    # syndication lenders attached
    r = await client.get(f"/v1/syndication/{tr['id']}/lenders")
    assert {ln["lender_name"] for ln in r.json()} == {"Axis Finance", "Bajaj Finance"}


async def test_merge_reimport_is_upsert_not_duplicate(client: AsyncClient):
    """A second (merge) import of the same workbook must not duplicate anything or trip a
    unique constraint — the reviewer's 'merge is not a real merge' finding. People,
    counterparties, leads, deals and trackers are reused by their natural key."""
    mis = _mini_mis()
    files = {"file": ("mis.xlsx", mis,
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = await client.post("/v1/import/atlas-xlsx",
                          params={"mode": "replace", "reason": "test import"}, files=files)
    assert r.status_code == 200, r.text

    ents = (await client.get("/v1/entities", params={"with_total": True})).json()["total"]
    leads = (await client.get("/v1/leads", params={"with_total": True})).json()["total"]
    deals = (await client.get("/v1/deals", params={"with_total": True})).json()["total"]
    syn = (await client.get("/v1/syndication", params={"with_total": True})).json()["items"]
    lenders = await client.get(f"/v1/syndication/{syn[0]['id']}/lenders")
    n_lenders = len(lenders.json())

    # Re-import the SAME workbook in merge mode — must be a no-op on counts, not a
    # constraint violation.
    files = {"file": ("mis.xlsx", mis,
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = await client.post("/v1/import/atlas-xlsx",
                          params={"mode": "merge", "reason": "test import"}, files=files)
    assert r.status_code == 200, r.text
    c = r.json()["counts"]
    assert c["entities"] == 0 and c["leads"] == 0 and c["deals"] == 0
    assert c["counterparties"] == 0 and c["people"] == 0
    assert c["syndication_tracker"] == 0 and c["syndication_lenders"] == 0

    assert (await client.get("/v1/entities", params={"with_total": True})).json()["total"] == ents
    assert (await client.get("/v1/leads", params={"with_total": True})).json()["total"] == leads
    assert (await client.get("/v1/deals", params={"with_total": True})).json()["total"] == deals
    syn2 = (await client.get("/v1/syndication", params={"with_total": True})).json()["items"]
    assert len(syn2) == len(syn)
    assert len((await client.get(f"/v1/syndication/{syn2[0]['id']}/lenders")).json()) == n_lenders


async def test_company_suffix_variants_canonicalise_to_one_entity(client: AsyncClient):
    """'Pvt Ltd', 'Private Limited' and 'Ltd' variants of the same name must resolve to a
    single entity — the reviewer's canonicalization finding."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Leads")
    ws.append(["Company Name", "RM Owner", "Sector"])
    ws.append(["Helios Power Private Limited", "Shubh Dave", "Solar"])
    ws = wb.create_sheet("Deals")
    ws.append(["Company Name", "RM", "Lending?", "Syndication?", "Asset Mon?", "Stage"])
    ws.append(["Helios Power Pvt Ltd", "Shubh Dave", "Yes", "No", "No", "In Screening"])
    ws = wb.create_sheet("Lending Tracker")
    ws.append(["Company Name", "Lending Amount (₹ Cr)", "RM"])
    ws.append(["Helios Power Ltd", 5.0, "Shubh Dave"])
    buf = io.BytesIO()
    wb.save(buf)

    files = {"file": ("mis.xlsx", buf.getvalue(),
                      "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    r = await client.post("/v1/import/atlas-xlsx",
                          params={"mode": "replace", "reason": "test import"}, files=files)
    assert r.status_code == 200, r.text
    # All three suffix variants collapse to ONE entity.
    assert r.json()["counts"]["entities"] == 1


async def test_import_rejects_non_xlsx(client: AsyncClient):
    files = {"file": ("data.txt", b"not a workbook", "text/plain")}
    r = await client.post("/v1/import/atlas-xlsx", params={"reason": "test import"}, files=files)
    assert r.status_code == 422


async def test_replace_import_is_tenant_scoped(client, tmp_path):
    """A replace import for tenant B must NOT delete tenant A's rows (the TRUNCATE
    data-loss fix). Seed a row on EVAM, run a replace import for a SECOND tenant, and
    confirm the EVAM row survives."""
    import uuid as _uuid

    from openpyxl import Workbook

    keep = (await client.post("/v1/entities", json={
        "code": f"KEEP-{_uuid.uuid4().hex[:6]}", "legal_name": "Keep Me Co"})).json()

    code = f"T2{_uuid.uuid4().hex[:4]}".upper()
    r = await client.post("/v1/tenants", json={"code": code, "name": "Second Tenant"})
    assert r.status_code in (201, 200), r.text

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append(["Company Name", "Sector", "RM Owner"])
    ws.append(["Second Tenant Co", "Solar - EPC", "Chetan"])
    path = tmp_path / "mis.xlsx"
    wb.save(path)
    with open(path, "rb") as fh:
        r = await client.post("/v1/import/atlas-xlsx?mode=replace&reason=test%20import",
                              files={"file": ("mis.xlsx", fh.read())},
                              headers={"X-Tenant": code})
    assert r.status_code == 200, r.text

    # The EVAM row is untouched — a TRUNCATE would have wiped it.
    still = await client.get(f"/v1/entities/{keep['id']}")
    assert still.status_code == 200


async def test_import_requires_a_reason(client, tmp_path):
    """A governed import bypasses the interactive lifecycle policy, so a non-empty reason is
    mandatory — an import without one is refused."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Leads"
    ws.append(["Company Name", "Sector"])
    ws.append(["Reason Co", "Solar - EPC"])
    path = tmp_path / "m.xlsx"
    wb.save(path)
    with open(path, "rb") as fh:
        r = await client.post("/v1/import/atlas-xlsx",
                              files={"file": ("m.xlsx", fh.read())})
    assert r.status_code == 422, r.text


async def test_import_quarantines_unknown_lifecycle_values_and_audits(client, tmp_path):
    """A spreadsheet row whose lifecycle value is unknown is QUARANTINED (skipped), surfaced in
    the response report, and the import is recorded in an immutable audit event with its file
    checksum, mode and reason."""
    from openpyxl import Workbook
    from sqlalchemy import text

    from app.db.session import get_sessionmaker
    wb = Workbook()
    leads = wb.active
    leads.title = "Leads"
    leads.append(["Company Name", "Sector"])
    leads.append(["QuarantineCo", "Solar - EPC"])
    lend = wb.create_sheet("Lending Tracker")
    lend.append(["Company Name", "Stage", "Lending Amount (₹ Cr)"])
    lend.append(["QuarantineCo", "Bogus Stage", 5])
    path = tmp_path / "q.xlsx"
    wb.save(path)
    with open(path, "rb") as fh:
        r = await client.post("/v1/import/atlas-xlsx",
                              params={"mode": "replace", "reason": "hist load", "ticket": "INC-9"},
                              files={"file": ("q.xlsx", fh.read())})
    assert r.status_code == 200, r.text
    body = r.json()
    # The bad lending row is quarantined, not imported.
    assert body["report"]["quarantined_count"] >= 1
    q = body["report"]["quarantined"][0]
    assert q["value"] == "Bogus Stage" and "unknown" in q["reason"].lower()
    assert body["counts"]["lending_tracker"] == 0
    assert len(body["checksum"]) == 64
    # An immutable audit event records the import with its checksum + reason + ticket.
    sm = get_sessionmaker()
    async with sm() as s:
        row = (await s.execute(text(
            "SELECT actor, changes FROM audit_log WHERE action='mis.import' "
            "AND resource_id=:c"), {"c": body["checksum"]})).first()
    assert row is not None
    assert row[1]["reason"] == "hist load" and row[1]["ticket"] == "INC-9"
    assert row[1]["quarantined_count"] >= 1


async def test_import_quarantines_incomplete_terminal_by_default(client, tmp_path):
    """A 'Ready for Disbursement' lending row missing its mandatory amount/date is QUARANTINED by
    default — the same state the interactive API rejects — and only imported (flagged
    reconciliation) under an explicit retain_incomplete override."""
    from openpyxl import Workbook

    def _wb():
        wb = Workbook()
        leads = wb.active
        leads.title = "Leads"
        leads.append(["Company Name", "Sector"])
        leads.append(["IncompleteCo", "Solar - EPC"])
        lend = wb.create_sheet("Lending Tracker")
        lend.append(["Company Name", "Stage"])
        lend.append(["IncompleteCo", "Ready for Disbursement"])   # no Disbursed Amount / Date
        p = tmp_path / "inc.xlsx"
        wb.save(p)
        return p

    with open(_wb(), "rb") as fh:
        r = await client.post("/v1/import/atlas-xlsx",
                              params={"mode": "replace", "reason": "load"},
                              files={"file": ("inc.xlsx", fh.read())})
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["report"]["quarantined_count"] >= 1
    assert body["counts"]["lending_tracker"] == 0
    # Explicit historical override → imported, flagged reconciliation_status=Required.
    with open(_wb(), "rb") as fh:
        r2 = await client.post("/v1/import/atlas-xlsx",
                               params={"mode": "replace", "reason": "hist",
                                       "retain_incomplete": "true"},
                               files={"file": ("inc.xlsx", fh.read())})
    b2 = r2.json()
    assert b2["report"]["reconciliation_count"] >= 1
    assert b2["report"]["reconciliation"][0]["reconciliation_status"] == "Required"
    assert "proposed_disbursement_amount" in b2["report"]["reconciliation"][0]["missing"]
    assert b2["counts"]["lending_tracker"] == 1
    assert b2["import_batch_id"]


async def test_merge_import_appends_stage_history(client, tmp_path):
    """A MERGE import that changes an existing tracker's stage appends an xlsx-import event to the
    append-only history (not a silent overwrite) and reports the change with a batch id."""
    from openpyxl import Workbook

    def _wb(stage):
        wb = Workbook()
        leads = wb.active
        leads.title = "Leads"
        leads.append(["Company Name", "Sector"])
        leads.append(["HistCo", "Solar - EPC"])
        lend = wb.create_sheet("Lending Tracker")
        lend.append(["Company Name", "Stage"])
        lend.append(["HistCo", stage])
        p = tmp_path / f"h_{stage}.xlsx"
        wb.save(p)
        return p

    with open(_wb("Diligence"), "rb") as fh:
        await client.post("/v1/import/atlas-xlsx", params={"mode": "replace", "reason": "a"},
                          files={"file": ("h.xlsx", fh.read())})
    with open(_wb("Note Circulated"), "rb") as fh:
        r = await client.post("/v1/import/atlas-xlsx", params={"mode": "merge", "reason": "b"},
                              files={"file": ("h.xlsx", fh.read())})
    body = r.json()
    assert body["report"]["history_change_count"] >= 1
    change = body["report"]["history_changes"][0]
    assert change["from"] == "Diligence" and change["to"] == "Note Circulated"
    # The tracker's own append-only history carries the xlsx-import event with the batch id.
    items = (await client.get("/v1/lending", params={"with_total": True})).json()["items"]
    hist = items[0]["stage_history"]
    assert hist and hist[-1]["source"] == "xlsx-import"
    assert hist[-1]["batch_id"] == body["import_batch_id"]


_ADMIN = {"X-User-Email": "admin@evamfinance.com", "X-User-Roles": "Admin"}
_MGMT = {"X-User-Email": "cro@evamfinance.com", "X-User-Roles": "Management"}


async def _import_incomplete_disbursed(client, tmp_path, company="ReconCo"):  # noqa: ANN001
    """Import a 'Ready for Disbursement' lending row missing its amount/date under the retain
    override; returns the reconciliation item and the lending subject id."""
    from openpyxl import Workbook
    wb = Workbook()
    leads = wb.active
    leads.title = "Leads"
    leads.append(["Company Name", "Sector"])
    leads.append([company, "Solar - EPC"])
    lend = wb.create_sheet("Lending Tracker")
    lend.append(["Company Name", "Stage"])
    lend.append([company, "Ready for Disbursement"])
    p = tmp_path / f"{company}.xlsx"
    wb.save(p)
    with open(p, "rb") as fh:
        r = await client.post("/v1/import/atlas-xlsx",
                              params={"mode": "replace", "reason": "hist",
                                      "retain_incomplete": "true"},
                              files={"file": (f"{company}.xlsx", fh.read())})
    assert r.status_code == 200, r.text
    items = (await client.get("/v1/reconciliation", params={"status": "Required"},
                              headers=_ADMIN)).json()["items"]
    item = next(i for i in items if i["company"] == company)
    return item, item["subject_id"]


async def test_incomplete_import_is_excluded_from_operational_reads(client, tmp_path):
    """An unresolved incomplete record must NOT appear in normal operational lists/totals — for
    services (fail closed) or ordinary reads — only via an Admin's explicit opt-in."""
    item, lid = await _import_incomplete_disbursed(client, tmp_path)
    assert item["subject_type"] == "Lending"
    assert "proposed_disbursement_amount" in item["missing_fields"] and item["original_values"]
    # Default (service) operational list EXCLUDES it — it can't count toward disbursed totals.
    default_list = (await client.get("/v1/lending", params={"with_total": True})).json()
    assert all(x["id"] != lid for x in default_list["items"])
    # A non-Admin opt-in is ignored (fail closed); an Admin opt-in reveals it, still flagged.
    admin_incl = (await client.get("/v1/lending",
                                   params={"include_reconciliation": "true"},
                                   headers=_ADMIN)).json()
    flagged = next(x for x in admin_incl["items"] if x["id"] == lid)
    assert flagged["reconciliation_status"] == "Required"


async def test_reconciliation_resolve_requires_actual_correction(client, tmp_path):
    """Resolving must PROVE the data is correct — a note alone is refused while a mandatory field
    is still missing; it succeeds only once the record is corrected, and then clears the flag."""
    from sqlalchemy import text

    from app.db.session import get_sessionmaker
    item, lid = await _import_incomplete_disbursed(client, tmp_path)
    # A non-Admin may not touch reconciliation.
    assert (await client.get("/v1/reconciliation")).status_code == 403
    # Resolve with only a note → refused (proposed_disbursement_amount/date still missing).
    bad = await client.post(f"/v1/reconciliation/{item['id']}/resolve",
                            json={"note": "looks fine"}, headers=_ADMIN)
    assert bad.status_code == 422, bad.text
    assert "still missing" in bad.text.lower()
    # Correct the record through its normal (policy-enforcing) API…
    fix = await client.patch(
        f"/v1/lending/{lid}",
        json={"proposed_disbursement_amount": 3.5, "proposed_disbursement_date": "2026-02-01"})
    assert fix.status_code == 200, fix.text
    # …now resolution succeeds, the flag clears, and the record re-enters operational reads.
    ok = await client.post(f"/v1/reconciliation/{item['id']}/resolve",
                           json={"note": "amount/date confirmed from sanction letter"},
                           headers=_ADMIN)
    assert ok.status_code == 200 and ok.json()["status"] == "Resolved"
    assert any(x["id"] == lid
               for x in (await client.get("/v1/lending")).json()["items"])
    sm = get_sessionmaker()
    async with sm() as s:
        row = (await s.execute(text(
            "SELECT changes FROM audit_log WHERE action='reconciliation.resolve' "
            "AND resource_id=:i"), {"i": item["id"]})).first()
    assert row is not None
    # The audit preserves the ORIGINAL incomplete import and the corrected (after) value.
    assert row[0]["original_values"].get("proposed_disbursement_amount") in (None, "")
    assert row[0]["after"]["proposed_disbursement_amount"] not in (None, "")


async def test_reconciliation_resolve_has_no_inline_write_bypass(client, tmp_path):
    """The resolve endpoint accepts NO business-field payload — a caller cannot slip an inline
    'corrected' (or stage/tenant/version) change through it, bypassing the update schema, policy
    engine and locks. Corrections must go through the record's own policy-enforcing API."""
    item, _lid = await _import_incomplete_disbursed(client, tmp_path, company="BypassCo")
    for body in (
        {"note": "x", "corrected": {"proposed_disbursement_amount": 9}},   # inline correction — forbidden
        {"note": "x", "corrected": {"stage": "Data Awaited"}},  # lifecycle jump — forbidden
        {"note": "x", "tenant_id": "00000000-0000-0000-0000-000000000000"},
        {"note": "x", "version": 99},
    ):
        r = await client.post(f"/v1/reconciliation/{item['id']}/resolve", json=body,
                              headers=_ADMIN)
        assert r.status_code == 422, f"{body} should be refused (extra fields forbidden): {r.text}"


async def test_reconciliation_waiver_requires_senior_authority(client, tmp_path):
    """A WAIVER keeps an incomplete record in the business of record — it is a senior business
    decision reserved to Management (a maker-checker style designated authority), NOT any single
    Admin operator who can otherwise run reconciliation."""
    item, _lid = await _import_incomplete_disbursed(client, tmp_path, company="AuthCo")
    # A plain Admin may list/assign/resolve, but may NOT waive.
    admin_waive = await client.post(f"/v1/reconciliation/{item['id']}/resolve",
                                    json={"status": "Waived", "note": "legacy", "ticket": "OPS-1"},
                                    headers=_ADMIN)
    assert admin_waive.status_code == 403, admin_waive.text
    assert "management" in admin_waive.text.lower()
    # Management can.
    ok = await client.post(f"/v1/reconciliation/{item['id']}/resolve",
                           json={"status": "Waived", "note": "legacy", "ticket": "OPS-1"},
                           headers=_MGMT)
    assert ok.status_code == 200 and ok.json()["status"] == "Waived"


async def test_reconciliation_waiver_requires_ticket_and_stays_marked(client, tmp_path):
    """Waived is a break-glass outcome keeping an incomplete record — it requires a ticket, and
    the subject stays visibly flagged 'Waived' (NOT cleared to look fully reconciled)."""
    item, lid = await _import_incomplete_disbursed(client, tmp_path, company="WaiveCo")
    no_ticket = await client.post(f"/v1/reconciliation/{item['id']}/resolve",
                                  json={"status": "Waived", "note": "legacy row, accept as-is"},
                                  headers=_MGMT)
    assert no_ticket.status_code == 422 and "ticket" in no_ticket.text.lower()
    ok = await client.post(f"/v1/reconciliation/{item['id']}/resolve",
                           json={"status": "Waived", "note": "legacy row", "ticket": "OPS-7"},
                           headers=_MGMT)
    assert ok.status_code == 200 and ok.json()["status"] == "Waived"
    # The subject is marked 'Waived' (a deliberate exception) — NOT cleared to None.
    lend = next(x for x in (await client.get("/v1/lending",
                                             params={"include_reconciliation": "true"},
                                             headers=_ADMIN)).json()["items"] if x["id"] == lid)
    assert lend["reconciliation_status"] == "Waived"


async def test_waived_record_is_excluded_from_default_operational_reads(client, tmp_path):
    """A WAIVED-but-incomplete record is a governed exception, NOT fully reconciled — it must stay
    OUT of routine operational lists/GET/exports/counts by default (so a waived incomplete line can
    never silently enter disbursed totals or trigger downstream), surfacing only under explicit
    Admin/Management inclusion."""
    item, lid = await _import_incomplete_disbursed(client, tmp_path, company="WaivedHidden")
    ok = await client.post(f"/v1/reconciliation/{item['id']}/resolve",
                           json={"status": "Waived", "note": "legacy row", "ticket": "OPS-9"},
                           headers=_MGMT)
    assert ok.status_code == 200 and ok.json()["status"] == "Waived"
    # Default (service) list, known-id GET, JSON export and counts all EXCLUDE the waived record.
    default_list = (await client.get("/v1/lending")).json()["items"]
    assert all(x["id"] != lid for x in default_list)
    assert (await client.get(f"/v1/lending/{lid}")).status_code == 404
    ex = (await client.get("/v1/export/json", params={"tables": "lending_tracker"})).json()
    assert all(row["id"] != lid for row in ex["tables"]["lending_tracker"])
    base_count = (await client.get("/v1/export/counts")).json().get("lending_tracker", 0)
    admin_count = (await client.get("/v1/export/counts",
                                    params={"include_reconciliation": "true"},
                                    headers=_ADMIN)).json().get("lending_tracker", 0)
    assert admin_count == base_count + 1
    # Under explicit Admin inclusion it is visible, still flagged 'Waived' (distinguishable).
    incl = next(x for x in (await client.get("/v1/lending",
                                             params={"include_reconciliation": "true"},
                                             headers=_ADMIN)).json()["items"] if x["id"] == lid)
    assert incl["reconciliation_status"] == "Waived"


async def test_concurrent_resolution_of_two_items_leaves_correct_flag(client, tmp_path):
    """Two admins resolving DIFFERENT items on the SAME subject at the same time must not race and
    leave the subject flag wrong. The item + subject + sibling-item FOR UPDATE locks serialize the
    two closing transactions, so the recompute always runs against the settled set — the final flag
    is deterministic (both resolved → cleared to None), never stuck 'Required'."""
    import asyncio

    from sqlalchemy import text

    from app.db.session import get_sessionmaker
    item1, lid = await _import_incomplete_disbursed(client, tmp_path, company="RaceCo")
    # Open a SECOND item on the same lending subject (a different missing field set).
    sm = get_sessionmaker()
    async with sm() as s:
        await s.execute(text(
            "INSERT INTO import_reconciliation_items "
            "(import_batch_id, subject_type, subject_id, missing_fields, status, tenant_id) "
            "SELECT 'batchB', 'Lending', CAST(:sid AS uuid), '[]'::jsonb, 'Required', "
            "tenant_id FROM lending_tracker WHERE id = CAST(:sid AS uuid)"), {"sid": lid})
        await s.commit()
    item2 = next(i for i in (await client.get("/v1/reconciliation", params={"status": "Required"},
                             headers=_ADMIN)).json()["items"]
                 if i["subject_id"] == lid and i["id"] != item1["id"])
    # Correct the record so BOTH items are resolvable (item2 flagged no fields).
    await client.patch(f"/v1/lending/{lid}",
                       json={"proposed_disbursement_amount": 4.0, "proposed_disbursement_date": "2026-04-01"})
    # Fire both resolutions concurrently — each runs in its own request/session/transaction.
    r1, r2 = await asyncio.gather(
        client.post(f"/v1/reconciliation/{item1['id']}/resolve",
                    json={"note": "amount/date filled"}, headers=_ADMIN),
        client.post(f"/v1/reconciliation/{item2['id']}/resolve",
                    json={"note": "nothing else missing"}, headers=_ADMIN))
    assert r1.status_code == 200 and r2.status_code == 200, (r1.text, r2.text)
    # With BOTH items resolved the flag must have cleared — a lost race would leave it 'Required'.
    lend = next(x for x in (await client.get("/v1/lending",
                                             params={"include_reconciliation": "true"},
                                             headers=_ADMIN)).json()["items"] if x["id"] == lid)
    assert lend["reconciliation_status"] is None, lend
    assert any(x["id"] == lid for x in (await client.get("/v1/lending")).json()["items"])


async def test_stale_if_match_version_is_refused(client, tmp_path):
    """Assignment/closure are idempotent via optimistic concurrency: a stale If-Match version is
    refused (409) so a caller working from an out-of-date view can't clobber a concurrent change."""
    item, _lid = await _import_incomplete_disbursed(client, tmp_path, company="IfMatchCo")
    ver = item["version"]
    # A matching version assigns fine and bumps the version.
    ok = await client.post(f"/v1/reconciliation/{item['id']}/assign",
                           json={"owner": "ops"}, headers={**_ADMIN, "If-Match": str(ver)})
    assert ok.status_code == 200, ok.text
    # Re-using the now-stale version is refused.
    stale = await client.post(f"/v1/reconciliation/{item['id']}/assign",
                              json={"owner": "ops2"}, headers={**_ADMIN, "If-Match": str(ver)})
    assert stale.status_code == 409, stale.text


async def test_incomplete_record_absent_from_get_export_and_counts(client, tmp_path):
    """An unresolved incomplete record must be invisible to a known-id GET, JSON export and counts
    for services (fail closed) — only an Admin opt-in reveals it."""
    item, lid = await _import_incomplete_disbursed(client, tmp_path, company="HiddenCo")
    # Known-id GET fails closed (404) for a service caller; Admin opt-in reveals it.
    assert (await client.get(f"/v1/lending/{lid}")).status_code == 404
    assert (await client.get(f"/v1/lending/{lid}",
                             params={"include_reconciliation": "true"},
                             headers=_ADMIN)).status_code == 200
    # JSON export (default) omits it; counts don't count it.
    ex = (await client.get("/v1/export/json", params={"tables": "lending_tracker"})).json()
    assert all(row["id"] != lid for row in ex["tables"]["lending_tracker"])
    base_count = (await client.get("/v1/export/counts")).json().get("lending_tracker", 0)
    admin_count = (await client.get("/v1/export/counts",
                                    params={"include_reconciliation": "true"},
                                    headers=_ADMIN)).json().get("lending_tracker", 0)
    assert admin_count == base_count + 1


async def test_flag_stays_until_all_items_for_a_subject_are_resolved(client, tmp_path):
    """A record with two open reconciliation items stays flagged until BOTH are resolved — one
    resolution must not make a still-incomplete record look reconciled."""
    from sqlalchemy import text

    from app.db.session import get_sessionmaker
    item, lid = await _import_incomplete_disbursed(client, tmp_path, company="TwoIssues")
    # Open a SECOND reconciliation item for the same lending subject.
    sm = get_sessionmaker()
    async with sm() as s:
        await s.execute(text(
            "INSERT INTO import_reconciliation_items "
            "(import_batch_id, subject_type, subject_id, missing_fields, status, tenant_id) "
            "SELECT 'batch2', 'Lending', CAST(:sid AS uuid), '[\"analyst\"]'::jsonb, 'Required', "
            "tenant_id FROM lending_tracker WHERE id = CAST(:sid AS uuid)"), {"sid": lid})
        await s.commit()
    # Correct + resolve the FIRST item.
    await client.patch(f"/v1/lending/{lid}",
                       json={"proposed_disbursement_amount": 2.0, "proposed_disbursement_date": "2026-03-01"})
    await client.post(f"/v1/reconciliation/{item['id']}/resolve",
                      json={"note": "amount/date filled"}, headers=_ADMIN)
    # The record is STILL flagged — the second issue remains open.
    still = (await client.get("/v1/lending", params={"include_reconciliation": "true"},
                              headers=_ADMIN)).json()["items"]
    assert next(x for x in still if x["id"] == lid)["reconciliation_status"] == "Required"


async def test_reconciliation_flag_is_scoped_by_subject_type(client, tmp_path):
    """Remaining-item matching keys on subject_type + subject_id — a Deal item that happens to
    share the Lending row's uuid (ids live in separate tables) must NOT keep the Lending flagged."""
    from sqlalchemy import text

    from app.db.session import get_sessionmaker
    item, lid = await _import_incomplete_disbursed(client, tmp_path, company="CrossType")
    sm = get_sessionmaker()
    async with sm() as s:
        await s.execute(text(
            "INSERT INTO import_reconciliation_items "
            "(import_batch_id, subject_type, subject_id, missing_fields, status, tenant_id) "
            "SELECT 'b', 'Deal', CAST(:sid AS uuid), '[\"rm\"]'::jsonb, 'Required', tenant_id "
            "FROM lending_tracker WHERE id = CAST(:sid AS uuid)"), {"sid": lid})
        await s.commit()
    # Fix + resolve the LENDING item — its flag clears despite the same-id Deal item still open.
    await client.patch(f"/v1/lending/{lid}",
                       json={"proposed_disbursement_amount": 1.0, "proposed_disbursement_date": "2026-01-01"})
    ok = await client.post(f"/v1/reconciliation/{item['id']}/resolve",
                           json={"note": "done"}, headers=_ADMIN)
    assert ok.status_code == 200, ok.text
    assert any(x["id"] == lid for x in (await client.get("/v1/lending")).json()["items"])


async def test_import_writes_initial_history_for_every_product_line(client, tmp_path):
    """A newly imported historical record created directly at an advanced stage gets an INITIAL
    null → stage history event — for EVERY product line (Deal, Lending, Syndication, AssetMon)."""
    from openpyxl import Workbook
    wb = Workbook()
    leads = wb.active
    leads.title = "Leads"
    leads.append(["Company Name", "Sector"])
    leads.append(["AllLines Co", "Solar - EPC"])
    deals = wb.create_sheet("Deals")
    deals.append(["Company Name", "Stage"])
    deals.append(["AllLines Co", "In Pipeline"])
    lend = wb.create_sheet("Lending Tracker")
    lend.append(["Company Name", "Stage"])
    lend.append(["AllLines Co", "Note Circulated"])
    syn = wb.create_sheet("Syndication")
    # The per-bank Status column drives the tracker's pipeline position (the coarse
    # Deal Status only overlays the Dropped/Disbursed terminals).
    syn.append(["Company Name", "Deal Status", "Bank", "Status"])
    syn.append(["AllLines Co", "Deal Live", "Axis Finance", "IM Circulated"])
    am = wb.create_sheet("Asset Mon")
    am.append(["Company Name", "Status"])
    am.append(["AllLines Co", "NBO Received"])
    p = tmp_path / "all.xlsx"
    wb.save(p)
    with open(p, "rb") as fh:
        r = await client.post("/v1/import/atlas-xlsx",
                              params={"mode": "replace", "reason": "historical load"},
                              files={"file": ("all.xlsx", fh.read())})
    assert r.status_code == 200, r.text

    deal = (await client.get("/v1/deals", params={"with_total": True})).json()["items"][0]
    assert deal["stage_history"][-1]["from"] is None and deal["stage_history"][-1]["to"] == "In Pipeline"
    assert deal["stage_history"][-1]["source"] == "xlsx-import"
    lend = (await client.get("/v1/lending", params={"with_total": True})).json()["items"][0]
    assert lend["stage_history"][-1]["to"] == "Note Circulated"
    syn = (await client.get("/v1/syndication", params={"with_total": True})).json()["items"][0]
    assert syn["status_history"][-1]["to"] == "IM Circulated"
    am = (await client.get("/v1/asset-monetisation", params={"with_total": True})).json()["items"][0]
    assert am["status_history"][-1]["to"] == "NBO Received"


async def test_legacy_stage_labels_map_to_new_vocabulary(client, tmp_path):
    """Legacy ATLAS 'Documentation' maps to 'CP/CS Completed'; 'Disbursed' imports VERBATIM,
    with its recorded amount/date becoming the proposed drawdown, so it is not quarantined for
    a missing proposed amount. A credit word on the DEALS sheet, by contrast, quarantines —
    the deal's stage is the funnel."""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    leads = wb.create_sheet("Leads")
    leads.append(["Company Name", "Sector"])
    leads.append(["LegacyCo", "Solar - EPC"])
    deals = wb.create_sheet("Deals")
    deals.append(["Company Name", "Stage"])
    # A credit-lifecycle word on the DEALS sheet is no longer deal vocabulary (the deal's stage
    # is the funnel) — this row must QUARANTINE by name, while the lending row still maps.
    deals.append(["LegacyCo", "Documentation"])
    lend = wb.create_sheet("Lending Tracker")
    lend.append(["Company Name", "Stage", "Disbursed Amount (₹ Cr)", "Disbursement Date"])
    lend.append(["LegacyCo", "Disbursed", 7.5, "2025-12-01"])
    p = tmp_path / "legacy.xlsx"
    wb.save(p)
    with open(p, "rb") as fh:
        r = await client.post("/v1/import/atlas-xlsx",
                              params={"mode": "replace", "reason": "legacy"},
                              files={"file": ("legacy.xlsx", fh.read())})
    assert r.status_code == 200, r.text
    assert r.json()["counts"]["lending_tracker"] == 1

    # The deal row was quarantined (credit words live on the Lending Tracker sheet now)…
    assert any(q["sheet"] == "Deals" and q["value"] == "Documentation"
               for q in r.json()["report"]["quarantined"])
    assert (await client.get("/v1/deals", params={"with_total": True})).json()["total"] == 0
    # …while the lending row still maps through the legacy-label aliases.
    row = (await client.get("/v1/lending", params={"with_total": True})).json()["items"][0]
    assert row["stage"] == "Disbursed"
    assert float(row["proposed_disbursement_amount"]) == 7.5


async def test_disbursed_without_disbursement_columns_derives_and_never_drops(client, tmp_path):
    """The live MIS Lending Tracker has NO disbursement columns — a 'Disbursed' facility carries
    only the facility amount and the stage-updated date. Those are DERIVED into the mandatory
    proposed amount/date (reported in the response's `derived` list) so the row imports; and a
    Disbursed row missing even those is a real exposure that imports FLAGGED for reconciliation
    instead of being dropped (the zero-omission rule)."""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    leads = wb.create_sheet("Leads")
    leads.append(["Company Name", "Sector"])
    leads.append(["MisShape Co", "Solar - EPC"])
    leads.append(["BareDisb Co", "Solar - EPC"])
    lend = wb.create_sheet("Lending Tracker")
    # The real sheet's exact columns: no disbursement, no proposed columns.
    lend.append(["Company Name", "Lending Amount (₹ Cr)", "RM", "Credit Analyst", "Stage",
                 "Stage Updated", "Remarks"])
    lend.append(["MisShape Co", 1.15, "Shubh Dave", "AT", "Disbursed", "2026-06-03", "Disbursed"])
    # …and one with not even an amount/date to derive from.
    lend.append(["BareDisb Co", None, "Shubh Dave", "AT", "Disbursed", None, "no data"])
    p = tmp_path / "mis_shape.xlsx"
    wb.save(p)
    with open(p, "rb") as fh:
        r = await client.post("/v1/import/atlas-xlsx",
                              params={"mode": "replace", "reason": "mis shape"},
                              files={"file": ("mis_shape.xlsx", fh.read())})
    assert r.status_code == 200, r.text
    body = r.json()
    # BOTH rows imported — nothing quarantined on the Lending sheet.
    assert body["counts"]["lending_tracker"] == 2
    assert not any(q["sheet"] == "Lending Tracker" for q in body["report"]["quarantined"])
    # The derivations are reported, never silent.
    der = {(d["company"], d["field"]): d for d in body["report"]["derived"]}
    assert der[("MisShape Co", "proposed_disbursement_amount")]["from_column"] == "Lending Amount (₹ Cr)"
    assert der[("MisShape Co", "proposed_disbursement_date")]["from_column"] == "Stage Updated"
    rows = {x["remarks"]: x for x in (await client.get(
        "/v1/lending", params={"with_total": True, "include_reconciliation": True},
        headers=_ADMIN)).json()["items"]}
    ok_row = rows["Disbursed"]
    assert ok_row["stage"] == "Disbursed"
    assert float(ok_row["proposed_disbursement_amount"]) == 1.15
    assert ok_row["proposed_disbursement_date"] == "2026-06-03"
    # The bare row imported FLAGGED (reconciliation Required), listed in the report.
    bare = rows["no data"]
    assert bare["reconciliation_status"] == "Required"
    assert any(x["company"] == "BareDisb Co" and "proposed_disbursement_amount" in x["missing"]
               for x in body["report"]["reconciliation"])


async def test_multiple_asset_mandates_per_company_all_import(client, tmp_path):
    """The MIS lists SEVERAL asset-sale mandates for one company (e.g. a 58MW sale, a 100MW
    land advisory and a dropped project). Each sheet row is its own asset_monetisation record —
    never blended into one row per company — and a merge re-import matches the company's rows
    in sheet order (updating in place, no duplicates)."""
    from openpyxl import Workbook

    def _wb(status_row2):  # noqa: ANN001
        wb = Workbook()
        wb.remove(wb.active)
        leads = wb.create_sheet("Leads")
        leads.append(["Company Name", "Sector"])
        leads.append(["MultiAsset Co", "Solar - EPC"])
        am = wb.create_sheet("Asset Mon")
        am.append(["Company Name", "Size (MW)", "Deal Type", "Status", "Notes"])
        am.append(["MultiAsset Co", 58, "Capital Market", "In Discussion", "Solar+BESS sale"])
        am.append(["MultiAsset Co", 100, "Project Advisory", status_row2, "land advisory"])
        am.append(["MultiAsset Co", 60, "Capital Market", "Dropped", "Mahagenco project"])
        p2 = tmp_path / "multi_asset.xlsx"
        wb.save(p2)
        return p2

    with open(_wb("Teaser Prepared"), "rb") as fh:
        r = await client.post("/v1/import/atlas-xlsx",
                              params={"mode": "replace", "reason": "multi-asset"},
                              files={"file": ("ma.xlsx", fh.read())})
    assert r.status_code == 200, r.text
    assert r.json()["counts"]["asset_monetisation"] == 3
    rows = (await client.get("/v1/asset-monetisation",
                             params={"with_total": True})).json()["items"]
    assert len(rows) == 3
    by_size = {float(x["size_mw"]): x for x in rows}
    assert by_size[58.0]["status"] == "In Discussion"
    assert by_size[100.0]["deal_type"] == "Project Advisory"
    assert by_size[60.0]["status"] == "Dropped"

    # Merge re-import with the 2nd mandate advanced → SAME three rows (in-order match), the
    # 2nd one updated, history recording the move.
    with open(_wb("Teaser Shared"), "rb") as fh:
        r = await client.post("/v1/import/atlas-xlsx",
                              params={"mode": "merge", "reason": "refresh"},
                              files={"file": ("ma.xlsx", fh.read())})
    assert r.status_code == 200, r.text
    c = r.json()["counts"]
    assert c["asset_monetisation"] == 0 and c["asset_monetisation_updated"] == 3
    rows = (await client.get("/v1/asset-monetisation",
                             params={"with_total": True})).json()["items"]
    assert len(rows) == 3
    adv = next(x for x in rows if float(x["size_mw"]) == 100.0)
    assert adv["status"] == "Teaser Shared"
    assert adv["status_history"][-1]["from"] == "Teaser Prepared"


async def test_multiple_lending_facilities_per_company_all_import(client, tmp_path):
    """A company holding TWO facilities gets two lending lines — one per sheet row."""
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    leads = wb.create_sheet("Leads")
    leads.append(["Company Name", "Sector"])
    leads.append(["TwoLoans Co", "Solar - EPC"])
    lend = wb.create_sheet("Lending Tracker")
    lend.append(["Company Name", "Lending Amount (₹ Cr)", "Stage"])
    lend.append(["TwoLoans Co", 5.0, "Diligence"])
    lend.append(["TwoLoans Co", 12.0, "Data Awaited"])
    p2 = tmp_path / "two_loans.xlsx"
    wb.save(p2)
    with open(p2, "rb") as fh:
        r = await client.post("/v1/import/atlas-xlsx",
                              params={"mode": "replace", "reason": "two loans"},
                              files={"file": ("tl.xlsx", fh.read())})
    assert r.status_code == 200, r.text
    assert r.json()["counts"]["lending_tracker"] == 2
    rows = (await client.get("/v1/lending", params={"with_total": True})).json()["items"]
    assert {float(x["amount_cr"]) for x in rows} == {5.0, 12.0}
    assert {x["stage"] for x in rows} == {"Diligence", "Data Awaited"}
